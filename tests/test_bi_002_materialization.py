import copy
import json
from io import BytesIO
from pathlib import Path
import unittest
from unittest.mock import patch

from openpyxl import load_workbook
from robo_dados_publicos.analytics.bi_materialization import (
    BIMaterializationError,
    build_manifest,
    build_plan,
    future_preflight,
    load_policy,
    plan_serving,
    render_xlsx,
    validate_future_root,
    validate_manifest,
    validate_t2_authorization,
)
from robo_dados_publicos.analytics.bi_model import load_contract

ROOT = Path(__file__).resolve().parents[1]
FIX = json.loads((ROOT / "tests/fixtures/bi_002_materialization_input.json").read_text())


class TestBI002(unittest.TestCase):
    def setUp(self):
        self.contract = load_contract()
        self.rows = FIX["datasets"]

    def stop(self, code, fn, *args, **kwargs):
        with self.assertRaisesRegex(BIMaterializationError, code):
            fn(*args, **kwargs)

    def test_contract_boundary_and_six_datasets(self):
        policy = load_policy()
        self.assertEqual(policy["future_drive_root"], "13_BI")
        self.assertEqual(len(policy["dataset_allowlist"]), 6)
        self.assertEqual(set(policy["dataset_allowlist"]), set(self.rows))
        self.assertEqual(FIX["classification"], "SYNTHETIC_SANITIZED_TEST_ONLY")
        auth = policy["authorization_contract"]
        self.assertTrue(auth["required"])
        self.assertFalse(auth["active_authorization_embedded"])
        self.assertEqual(auth["tier"], "T2_CREATE_ONLY")
        self.assertEqual(auth["drive_root"], "13_BI")

    def test_reserved_roots(self):
        for root in load_policy()["reserved_roots"]:
            self.stop("STOP_BI_RESERVED_ROOT", validate_future_root, root)
        self.stop("STOP_BI_RESERVED_ROOT", validate_future_root, "09_SCRIPTS")

    def test_all_datasets_and_xlsx_semantics(self):
        for dataset, rows in self.rows.items():
            plan = build_plan(dataset, rows)
            data = render_xlsx(plan)
            wb = load_workbook(BytesIO(data), data_only=False)
            self.assertEqual(wb.sheetnames, [dataset[:31]])
            ws = wb.active
            self.assertEqual(ws.max_row, plan.row_count + 1)
            self.assertEqual(ws.max_column, len(plan.ordered_columns))
            self.assertEqual(tuple(c.value for c in ws[1]), plan.ordered_columns)
            self.assertFalse(ws.merged_cells)
            self.assertTrue(all(c.data_type != "f" for row in ws for c in row))
            manifest = build_manifest(plan, data)
            self.assertEqual(
                validate_manifest(plan, manifest, data),
                "PASS_BI_MATERIALIZATION_PLAN_OFFLINE",
            )
            self.assertEqual(
                manifest["schema_fingerprint_sha256"],
                plan.schema_fingerprint_sha256,
            )

    def test_order_hash_and_filename_determinism(self):
        row = self.rows["BI_SIOPE_SERIES"][0]
        other = {
            **row,
            "year": 2017,
            "annual_period": "P6",
            "source_sha256": "c" * 64,
        }
        a = build_plan("BI_SIOPE_SERIES", [other, row])
        b = build_plan("BI_SIOPE_SERIES", [row, other])
        self.assertEqual(
            (
                a.rows,
                a.schema_fingerprint_sha256,
                a.canonical_matrix_sha256,
                a.snapshot_id,
                a.proposed_snapshot_filename,
            ),
            (
                b.rows,
                b.schema_fingerprint_sha256,
                b.canonical_matrix_sha256,
                b.snapshot_id,
                b.proposed_snapshot_filename,
            ),
        )
        changed = {**other, "metric_value": 999}
        self.assertNotEqual(
            a.canonical_matrix_sha256,
            build_plan("BI_SIOPE_SERIES", [row, changed]).canonical_matrix_sha256,
        )
        self.assertEqual(render_xlsx(a), render_xlsx(b))

    def test_schema_semantics_are_part_of_snapshot_identity(self):
        journal = build_plan(
            "BI_JORNAL_EVENTOS", self.rows["BI_JORNAL_EVENTOS"], self.contract
        )
        matrix = json.loads(journal.canonical_matrix)
        fields = matrix["schema"]["fields"]
        positions = {field["name"]: index for index, field in enumerate(fields)}
        self.assertEqual(matrix["rows"][0][positions["publication_date"]][0], "date")
        self.assertEqual(matrix["rows"][0][positions["value"]][0], "currency")

        changed_contract = copy.deepcopy(self.contract)
        journal_spec = next(
            item
            for item in changed_contract["datasets"]
            if item["dataset_id"] == "BI_JORNAL_EVENTOS"
        )
        value_field = next(
            field for field in journal_spec["fields"] if field["name"] == "value"
        )
        value_field["data_type"] = "number"
        changed = build_plan(
            "BI_JORNAL_EVENTOS",
            self.rows["BI_JORNAL_EVENTOS"],
            changed_contract,
        )
        self.assertNotEqual(
            journal.schema_fingerprint_sha256, changed.schema_fingerprint_sha256
        )
        self.assertNotEqual(
            journal.canonical_matrix_sha256, changed.canonical_matrix_sha256
        )
        self.assertNotEqual(journal.snapshot_id, changed.snapshot_id)

    def test_duplicate_unknown_privacy_fail_closed(self):
        dataset = "BI_JORNAL_EVENTOS"
        row = self.rows[dataset][0]
        self.stop(
            "STOP_BI_DUPLICATE_PRIMARY_KEY", build_plan, dataset, [row, row]
        )
        self.stop(
            "STOP_BI_INVALID_SCHEMA", build_plan, dataset, [{**row, "invented": 1}]
        )
        self.stop(
            "STOP_BI_INVALID_SCHEMA",
            build_plan,
            dataset,
            [{**row, "refresh_token": "x"}],
        )
        self.stop("STOP_BI_UNKNOWN_DATASET", build_plan, "BI_ALERTAS", [])

    def test_siope_boundaries(self):
        row = self.rows["BI_SIOPE_SERIES"][0]
        self.assertEqual(build_plan("BI_SIOPE_SERIES", [row]).row_count, 1)
        for values in (
            {"annual_period": "P6"},
            {"year": 2025, "annual_period": "P6"},
            {"year": 2017, "annual_period": "P1"},
        ):
            self.stop(
                "STOP_BI_INVALID_SCHEMA",
                build_plan,
                "BI_SIOPE_SERIES",
                [{**row, **values}],
            )
        for year in range(2017, 2025):
            self.assertEqual(
                build_plan(
                    "BI_SIOPE_SERIES",
                    [{**row, "year": year, "annual_period": "P6"}],
                ).row_count,
                1,
            )

    def test_candidate_never_identity(self):
        row = self.rows["BI_RECONCILIACAO"][0]
        self.assertFalse(build_plan("BI_RECONCILIACAO", [row]).rows[0][14])
        self.stop(
            "STOP_BI_INVALID_SCHEMA",
            build_plan,
            "BI_RECONCILIACAO",
            [{**row, "financial_identity_proven": True}],
        )

    def test_manifest_mismatch_and_snapshot_mismatch(self):
        plan = build_plan("BI_DICIONARIO", self.rows["BI_DICIONARIO"])
        xlsx = render_xlsx(plan)
        manifest = build_manifest(plan, xlsx)
        self.stop(
            "STOP_BI_MANIFEST_MISMATCH",
            validate_manifest,
            plan,
            {**manifest, "row_count": 99},
            xlsx,
        )
        self.stop(
            "STOP_BI_SNAPSHOT_ID_MISMATCH",
            future_preflight,
            plan,
            {**manifest, "snapshot_id": "bad"},
        )
        self.stop(
            "STOP_BI_MANIFEST_MISMATCH",
            future_preflight,
            plan,
            {**manifest, "schema_fingerprint_sha256": "0" * 64},
        )

    def _authorization_shape(self, **updates):
        value = {
            "authorization_id": "UNIT_TEST_SHAPE_ONLY_NEVER_OPERATIONAL",
            "authorized": True,
            "repository": "ferinbon-cpu/robo-dados-publicos",
            "tier": "T2_CREATE_ONLY",
            "drive_root": "13_BI",
            "task": "BI_002_FIRST_T2_CREATE_ONLY_MATERIALIZATION",
            "scope": "13_BI_CREATE_ONLY_SNAPSHOTS",
            "implementation_sha": "a" * 40,
            "consumed": False,
            "test_only": False,
        }
        value.update(updates)
        return value

    def test_t2_authorization_is_sha_and_scope_bound(self):
        self.stop(
            "STOP_BI_T2_NOT_AUTHORIZED",
            validate_t2_authorization,
            None,
            expected_implementation_sha="a" * 40,
        )
        self.stop(
            "STOP_BI_T2_AUTHORIZATION_TEST_ONLY",
            validate_t2_authorization,
            self._authorization_shape(test_only=True),
            expected_implementation_sha="a" * 40,
        )
        self.stop(
            "STOP_BI_T2_AUTHORIZATION_CONSUMED",
            validate_t2_authorization,
            self._authorization_shape(consumed=True),
            expected_implementation_sha="a" * 40,
        )
        self.stop(
            "STOP_BI_T2_AUTHORIZATION_IMPLEMENTATION_SHA_INVALID",
            validate_t2_authorization,
            self._authorization_shape(),
            expected_implementation_sha="not-a-sha",
        )
        self.stop(
            "STOP_BI_T2_AUTHORIZATION_IMPLEMENTATION_SHA_MISMATCH",
            validate_t2_authorization,
            self._authorization_shape(implementation_sha="b" * 40),
            expected_implementation_sha="a" * 40,
        )
        for update in (
            {"repository": "other/repo"},
            {"tier": "T3_MUTATING_OR_PUBLICATION"},
            {"drive_root": "08_OUTPUTS"},
            {"task": "TASK_018"},
            {"scope": "TASK_018_REUSE"},
        ):
            self.stop(
                "STOP_BI_T2_AUTHORIZATION_SCOPE_MISMATCH",
                validate_t2_authorization,
                self._authorization_shape(**update),
                expected_implementation_sha="a" * 40,
            )

    def test_future_boundaries(self):
        plan = build_plan("BI_DICIONARIO", self.rows["BI_DICIONARIO"])
        manifest = build_manifest(plan, render_xlsx(plan))
        self.stop(
            "STOP_BI_REMOTE_COLLISION_REQUIRES_READBACK",
            future_preflight,
            plan,
            manifest,
            remote_collision=True,
        )
        self.stop(
            "STOP_BI_T2_NOT_AUTHORIZED",
            future_preflight,
            plan,
            manifest,
            expected_implementation_sha="a" * 40,
        )
        self.stop(
            "STOP_BI_T2_AUTHORIZATION_IMPLEMENTATION_SHA_MISMATCH",
            future_preflight,
            plan,
            manifest,
            authorization=self._authorization_shape(implementation_sha="b" * 40),
            expected_implementation_sha="a" * 40,
        )
        self.stop(
            "STOP_BI_SERVING_MUTATION_NOT_AUTHORIZED",
            plan_serving,
            snapshot_validated=True,
        )
        self.assertEqual(
            plan_serving(snapshot_validated=True, t3_authorized=True),
            "PASS_BI_MATERIALIZATION_PLAN_OFFLINE",
        )
        self.stop(
            "STOP_BI_LOOKER_SEPARATE_AUTHORIZATION_REQUIRED",
            plan_serving,
            snapshot_validated=True,
            t3_authorized=True,
            looker_authorized=True,
        )

    def test_zero_network_transport_publication(self):
        import robo_dados_publicos.analytics.bi_materialization as mod

        source = Path(mod.__file__).read_text()
        forbidden = (
            "requests",
            "googleapiclient",
            "socket",
            "DriveService",
            "schedule",
            "recurrence",
        )
        self.assertTrue(all(word not in source for word in forbidden))
        policy = load_policy()
        self.assertFalse(policy["remote_execution_authorized"])
        self.assertFalse(policy["authorization_contract"]["active_authorization_embedded"])
        with patch("socket.socket", side_effect=AssertionError("network")):
            build_plan("BI_DICIONARIO", self.rows["BI_DICIONARIO"])


if __name__ == "__main__":
    unittest.main()
