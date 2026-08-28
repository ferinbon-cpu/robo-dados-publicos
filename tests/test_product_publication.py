import csv
import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from robo_dados_publicos.core.models import AnswerContract
from robo_dados_publicos.product import build_product_report, write_product_bundle
from robo_dados_publicos.product.publication import (
    CSV_MIME,
    GOOGLE_SHEETS_MIME,
    XLSX_MIME,
    ProductPublicationError,
    publish_product_bundle,
)


class FakeDrive:
    def __init__(self, *, children=None, import_formats=None, fail_on=None, sheet_export_matrix=None):
        self.children = list(children or [])
        self.formats = import_formats if import_formats is not None else {XLSX_MIME: [GOOGLE_SHEETS_MIME]}
        self.fail_on = fail_on
        self.sheet_export_matrix = sheet_export_matrix
        self.calls = []
        self.meta = {}
        self.uploaded_payloads = {}
        self.imported_sheet_matrix = None

    def import_formats(self):
        self.calls.append(("import_formats",))
        return self.formats

    def list_children(self, parent_id):
        self.calls.append(("list_children", parent_id))
        return list(self.children)

    def put_converted(self, local_path, remote_name, parent_id, source_mime_type, target_mime_type):
        self.calls.append(("put_converted", remote_name, source_mime_type, target_mime_type))
        if self.fail_on == "sheet":
            raise RuntimeError("synthetic sheet failure")
        if source_mime_type != XLSX_MIME:
            raise RuntimeError("unexpected non-XLSX source")
        workbook = load_workbook(local_path, read_only=True, data_only=True)
        sheet = workbook.active
        self.imported_sheet_matrix = [
            ["" if value is None else str(value) for value in row]
            for row in sheet.iter_rows(values_only=True)
        ]
        workbook.close()
        self.meta["S1"] = {
            "id": "S1",
            "name": remote_name,
            "mimeType": target_mime_type,
            "parents": [parent_id],
        }
        return {"id": "S1", "name": remote_name, "mimeType": target_mime_type, "parents": [parent_id]}

    def export(self, file_id, destination, mime_type):
        self.calls.append(("export", file_id, mime_type))
        if self.fail_on == "sheet_export":
            raise RuntimeError("synthetic export failure")
        if mime_type != CSV_MIME:
            raise RuntimeError("unexpected export mime")
        matrix = self.sheet_export_matrix if self.sheet_export_matrix is not None else self.imported_sheet_matrix
        with Path(destination).open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle, lineterminator="\n")
            writer.writerows(matrix)
        return {"file_id": file_id, "path": str(destination), "mime_type": mime_type}

    def put(self, local_path, remote_name, parent_id, mime_type):
        kind = "manifest" if remote_name.endswith("_publication_manifest.json") else "pdf"
        self.calls.append(("put", kind, remote_name))
        if self.fail_on == kind:
            raise RuntimeError(f"synthetic {kind} failure")
        file_id = "M1" if kind == "manifest" else "P1"
        payload = Path(local_path).read_bytes()
        self.uploaded_payloads[remote_name] = payload
        self.meta[file_id] = {
            "id": file_id,
            "name": remote_name,
            "mimeType": mime_type,
            "size": str(len(payload)),
            "parents": [parent_id],
        }
        return {"id": file_id, "name": remote_name, "mimeType": mime_type, "parents": [parent_id]}

    def metadata(self, file_id):
        self.calls.append(("metadata", file_id))
        return dict(self.meta[file_id])


class TestProductPublication(unittest.TestCase):
    def make_bundle(self, root: Path) -> Path:
        answer = AnswerContract(
            status="ANSWERED",
            dado="Gate técnico aprovado.",
            calculo="Sem cálculo substantivo.",
            correspondencia="Não aplicável.",
            interpretacao="Teste de publicação controlada.",
            cautela="RELATÓRIO DE VALIDAÇÃO TÉCNICA.",
            fontes=("https://example.test/pr/15",),
        )
        report = build_product_report(
            [answer],
            report_id="M6_PRODUCT_OUTPUT_GATE_0_7_0",
            title="Gate técnico",
            scope="08_OUTPUTS",
            generated_at="2026-08-24T22:45:00+00:00",
        )
        out = root / "bundle"
        write_product_bundle(report, out)
        return out

    def publish(self, drive, bundle):
        return publish_product_bundle(
            drive,
            output_parent_id="OUTPUTS_PARENT",
            bundle_dir=bundle,
            remote_basename="ROBO_DADOS_PUBLICOS_M6_GATE_0_7_0",
            expected_report_status="READY_WITH_CAUTION",
            gate_id="M6_FIRST_PRODUCT_OUTPUT_PUBLICATION_GATE_0_7_0",
            published_at="2026-08-24T22:50:00+00:00",
        )

    def test_success_creates_exactly_sheet_pdf_and_manifest_in_order(self):
        with tempfile.TemporaryDirectory() as raw:
            drive = FakeDrive()
            result = self.publish(drive, self.make_bundle(Path(raw)))
        self.assertEqual("PASS_M6_PRODUCT_OUTPUT_PUBLICATION_GATE", result["status"])
        self.assertEqual(3, result["created_count"])
        writes = [call for call in drive.calls if call[0] in {"put", "put_converted"}]
        self.assertEqual("put_converted", writes[0][0])
        self.assertEqual(XLSX_MIME, writes[0][2])
        self.assertEqual("pdf", writes[1][1])
        self.assertEqual("manifest", writes[2][1])
        self.assertTrue(result["sheet_semantic_readback_verified"])
        self.assertEqual("XLSX_LOCALE_INDEPENDENT", result["sheet_import_transport"])
        self.assertGreaterEqual(result["sheet_rows"], 2)
        self.assertEqual(7, result["sheet_columns"])
        self.assertTrue(any(call[0] == "export" for call in drive.calls))
        self.assertTrue(result["completion_manifest_written_last"])
        self.assertFalse(result["overwrite_performed"])
        self.assertFalse(result["remote_identifiers_exposed"])

    def test_collision_stops_before_any_write(self):
        drive = FakeDrive(children=[{"name": "ROBO_DADOS_PUBLICOS_M6_GATE_0_7_0.pdf"}])
        with tempfile.TemporaryDirectory() as raw:
            with self.assertRaisesRegex(ProductPublicationError, "STOP_PRODUCT_OUTPUT_NAME_COLLISION"):
                self.publish(drive, self.make_bundle(Path(raw)))
        writes = [call for call in drive.calls if call[0] in {"put", "put_converted"}]
        self.assertEqual([], writes)

    def test_unsupported_xlsx_import_stops_before_inventory_and_writes(self):
        drive = FakeDrive(import_formats={XLSX_MIME: []})
        with tempfile.TemporaryDirectory() as raw:
            with self.assertRaisesRegex(ProductPublicationError, "STOP_PRODUCT_XLSX_TO_SHEETS_NOT_SUPPORTED"):
                self.publish(drive, self.make_bundle(Path(raw)))
        self.assertEqual([("import_formats",)], drive.calls)

    def test_tampered_bundle_stops_before_any_remote_request(self):
        drive = FakeDrive()
        with tempfile.TemporaryDirectory() as raw:
            bundle = self.make_bundle(Path(raw))
            with (bundle / "report.pdf").open("ab") as f:
                f.write(b"tamper")
            with self.assertRaisesRegex(ProductPublicationError, "STOP_PRODUCT_BUNDLE_SIZE_MISMATCH"):
                self.publish(drive, bundle)
        self.assertEqual([], drive.calls)

    def test_wrong_expected_status_stops_before_remote_request(self):
        drive = FakeDrive()
        with tempfile.TemporaryDirectory() as raw:
            bundle = self.make_bundle(Path(raw))
            with self.assertRaisesRegex(ProductPublicationError, "STOP_PRODUCT_REPORT_STATUS_MISMATCH"):
                publish_product_bundle(
                    drive,
                    output_parent_id="OUTPUTS_PARENT",
                    bundle_dir=bundle,
                    remote_basename="ROBO_DADOS_PUBLICOS_M6_GATE_0_7_0",
                    expected_report_status="READY",
                    gate_id="M6_FIRST_PRODUCT_OUTPUT_PUBLICATION_GATE_0_7_0",
                    published_at="2026-08-24T22:50:00+00:00",
                )
        self.assertEqual([], drive.calls)

    def test_completion_manifest_contains_no_remote_ids_and_records_hardening(self):
        with tempfile.TemporaryDirectory() as raw:
            drive = FakeDrive()
            self.publish(drive, self.make_bundle(Path(raw)))
        name = "ROBO_DADOS_PUBLICOS_M6_GATE_0_7_0_publication_manifest.json"
        text = drive.uploaded_payloads[name].decode("utf-8")
        payload = json.loads(text)
        self.assertNotIn("S1", text)
        self.assertNotIn("P1", text)
        self.assertNotIn("M1", text)
        self.assertFalse(payload["remote_identifiers_recorded"])
        self.assertTrue(payload["completion_marker_written_last"])
        self.assertEqual("XLSX_LOCALE_INDEPENDENT", payload["sheet_import_transport"])
        self.assertTrue(payload["sheet_semantic_readback_required"])

    def test_remote_failure_reports_created_count_without_automatic_cleanup(self):
        drive = FakeDrive(fail_on="pdf")
        with tempfile.TemporaryDirectory() as raw:
            with self.assertRaises(ProductPublicationError) as ctx:
                self.publish(drive, self.make_bundle(Path(raw)))
        self.assertEqual("STOP_PRODUCT_PUBLICATION_REMOTE_OPERATION", ctx.exception.code)
        self.assertEqual(1, ctx.exception.created_count)
        self.assertFalse(any(call[0] == "put" and call[1] == "manifest" for call in drive.calls))

    def test_sheet_export_failure_stops_after_sheet_and_before_pdf_manifest(self):
        drive = FakeDrive(fail_on="sheet_export")
        with tempfile.TemporaryDirectory() as raw:
            with self.assertRaises(ProductPublicationError) as ctx:
                self.publish(drive, self.make_bundle(Path(raw)))
        self.assertEqual("STOP_PRODUCT_SHEET_READBACK_EXPORT", ctx.exception.code)
        self.assertEqual(1, ctx.exception.created_count)
        self.assertFalse(any(call[0] == "put" for call in drive.calls))

    def test_sheet_content_mismatch_stops_after_sheet_and_before_pdf_manifest(self):
        drive = FakeDrive(sheet_export_matrix=[["status,DADO,CÁLCULO,CORRESPONDÊNCIA,INTERPRETAÇÃO,CAUTELA,FONTES"]])
        with tempfile.TemporaryDirectory() as raw:
            with self.assertRaises(ProductPublicationError) as ctx:
                self.publish(drive, self.make_bundle(Path(raw)))
        self.assertEqual("STOP_PRODUCT_SHEET_CONTENT_VERIFY", ctx.exception.code)
        self.assertEqual(1, ctx.exception.created_count)
        self.assertFalse(any(call[0] == "put" for call in drive.calls))

    def test_invalid_basename_stops_before_remote_request(self):
        drive = FakeDrive()
        with tempfile.TemporaryDirectory() as raw:
            bundle = self.make_bundle(Path(raw))
            with self.assertRaisesRegex(ProductPublicationError, "BASENAME_INVALID"):
                publish_product_bundle(
                    drive,
                    output_parent_id="OUTPUTS_PARENT",
                    bundle_dir=bundle,
                    remote_basename="../unsafe",
                    expected_report_status="READY_WITH_CAUTION",
                    gate_id="GATE",
                    published_at="2026-08-24T22:50:00+00:00",
                )
        self.assertEqual([], drive.calls)


if __name__ == "__main__":
    unittest.main()
